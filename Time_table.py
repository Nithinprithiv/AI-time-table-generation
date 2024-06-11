import kivy
from kivymd.app import MDApp
from kivy.uix.label import Label as L
from kivy.uix.button import Button as B
from kivy.uix.checkbox import CheckBox as C
from kivy.uix.image import Image as I
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.lang import Builder
import openpyxl as xl
from openpyxl import Workbook
import random as ra


class Manager(ScreenManager):
    Builder.load_string("""
<Manager>:
    Screen:
        name:'logo'
        canvas:
            Rectangle:
                pos: self.pos
                size: self.size
                source:'image/1.jpg'
        Label:
            text:"AI-TIME TABLE "
            color: 1,1,1,1
            font_style:"bold"
            font_size: 90
            pos_hint:{"center_x":0.5,"center_y":0.7}
        Label:
            text:"Generator"
            color: 1,1,1,1
            font_style:"italic"
            font_size: 40
            pos_hint:{"center_x":0.51,"center_y":0.63}
        
        MDRectangleFlatButton:
            text: "Next.."
            pos_hint:{"center_x":0.51,"center_y":0.5}
            theme_text_color: "Custom"
            text_color: "white"
            line_color: "green"
            on_press:
                root.current='user'
        Label:
            id:label
            text:""
            color: 1,0,0,1
            font_style:"italic"
            font_size: 20
            pos_hint:{"center_x":0.51,"center_y":0.25}
    Screen:
        name:'user'
        canvas:
            Rectangle:
                pos: self.pos
                size: self.size
                source:'image/user.jpg'
        Label:
            text:"No.Of.Class To Gernerate Time Table"
            color: 1,1,1,1
            font_style:"italic"
            font_size: 30
            pos_hint:{"center_x":0.5,"center_y":0.7}
        MDTextField:
            id:count
            mode: "rectangle"
            hint_text: "count"
            multiline:False
            color_mode: 'custom'
            line_color_focus: 1, 1, 1, 1
            size_hint:.3,.1
            pos_hint:{"center_x":0.5,"center_y":0.6}
        Label:
            text:"No.Of.Periods :"
            color: 1,1,1,1
            font_style:"italic"
            font_size: 30
            pos_hint:{"center_x":0.3,"center_y":0.5}
        MDTextField:
            id:period
            mode: "rectangle"
            hint_text: "period"
            multiline:False
            color_mode: 'custom'
            line_color_focus: 1, 1, 1, 1
            size_hint:.3,.1
            pos_hint:{"center_x":0.6,"center_y":0.5}
        Label:
            text:"No.Of.Days :"
            color: 1,1,1,1
            font_style:"italic"
            font_size: 30
            pos_hint:{"center_x":0.3,"center_y":0.4}
        MDTextField:
            id:day
            mode: "rectangle"
            hint_text: "days"
            multiline:False
            color_mode: 'custom'
            line_color_focus: 1, 1, 1, 1
            size_hint:.3,.1
            pos_hint:{"center_x":0.6,"center_y":0.4}
        MDRectangleFlatButton:
            text: "Generate input file..>"
            pos_hint:{"center_x":0.5,"center_y":0.3}
            theme_text_color: "Custom"
            text_color: "white"
            line_color: "green"
            on_press:
                root.current='gen'  
                app.input()                    
                
    Screen:
        name:'gen'
        canvas:
            Rectangle:
                pos: self.pos
                size: self.size
                source:'image/user.jpg'
        Label:
            text:"Input_Analytic_File Created "
            color: 0,1,1,1
            font_style:"italic"
            font_size: 40
            pos_hint:{"center_x":0.5,"center_y":0.7}
        Label:
            text:"Upload the Details About Subjects ..>>"
            color: 1,1,1,1
            font_style:"italic"
            font_size: 20
            pos_hint:{"center_x":0.5,"center_y":0.6}
        
        MDRectangleFlatButton:
            text: "Generate.."
            pos_hint:{"center_x":0.5,"center_y":0.5}
            theme_text_color: "Custom"
            text_color: "white"
            line_color: "green"
            on_press:
                root.current='gen1' 
                app.generate()
    Screen:
        name:'gen1'
        canvas:
            Rectangle:
                pos: self.pos
                size: self.size
                source:'image/user.jpg'
        Label:
            text:"Time Table Generated"
            color: 0,1,1,1
            font_style:"italic"
            font_size: 40
            pos_hint:{"center_x":0.5,"center_y":0.7}
        Label:
            text:"If you want Re-Generate..>"
            color: 1,1,1,1
            font_style:"italic"
            font_size: 20
            pos_hint:{"center_x":0.5,"center_y":0.6}
        
        MDRectangleFlatButton:
            text: "Re-Generate.."
            pos_hint:{"center_x":0.5,"center_y":0.5}
            theme_text_color: "Custom"
            text_color: "white"
            line_color: "green"
            on_press:
                app.generate() 
        MDRectangleFlatButton:
            text: "Exit"
            pos_hint:{"center_x":0.5,"center_y":0.4}
            theme_text_color: "Custom"
            text_color: "white"
            line_color: "red"
            on_press:
                root.current='logo'    
    """)
    pass
class AI_TimeTable(MDApp):
    def build(self):
        self.theme_cls.theme_style = "Dark"
        self.theme_cls.primary_palette = "Blue"
        return Manager()
    def input(self):
        m=self.root.ids.count.text
        n=int(m)
        w=Workbook()
        sheet1=w.active
        sheet_name=['sheet']

        sheet1["B2"] = "Faculty Name"
        sheet1["C2"] = "subject Name"
        sheet1["D2"] = "Acronym"
        sheet1["E2"] = "weekly period"

        for i in range(1,n):
            worksheet=w.create_sheet(title='sheet '+(str(i)))
            sheet_name.append('sheet '+(str(i)))
            worksheet.cell(row=2,column=2).value='Faculty Name'
            worksheet.cell(row=2,column=3).value='subject Name'
            worksheet.cell(row=2,column=4).value='Acronym'
            worksheet.cell(row=2,column=5).value='weekly period'
        sheet = w['sheet 1']
        print('Completed...')
        w.save('Input_Analytic.xlsx')
    def generate(self):
        ins=xl.load_workbook('Input_Analytic.xlsx')
        ins_sheet=ins.active

        # print(ins['sheet 1'].cell(row=2,column=3).value)
        sub=[];facalty=[];acr=[];priod=[];f=[];s=[];a=[];p=[]
        #
        for col in range(2,6):
                    for name in ins.sheetnames:
                        ins_s=ins[name]
                        for row in range(3,(ins[name].max_row)+1):
                            if col==2:
                                f.append((ins[name].cell(row=row,column=col).value))
                            elif col==3:
                                s.append((ins[name].cell(row=row,column=col).value))
                            elif col==4:
                                a.append((ins[name].cell(row=row,column=col).value))
                            elif col==5:
                                p.append((ins[name].cell(row=row,column=col).value))
                        if col==2:
                                facalty.append(f.copy())
                                f.clear()
                        elif col==3:
                                sub.append(s.copy())
                                s.clear()
                        elif col==4:
                                acr.append(a.copy())
                                a.clear()
                        elif col==5:
                                priod.append(p.copy())
                                p.clear()
        f.extend(facalty.copy())

        # print(sub.index([0],'CNS'))
        dic={}
        up={}
        for i in range(len(facalty)):
            for j in range(len(facalty[i])):
                dic[acr[i][j]]=facalty[i][j]
                up[acr[i][j]]=priod[i][j]
        lab1=[]
        lab=[]
        for i in range(len(acr)):
            for j in range(0,len(acr[i])):
                if acr[i][j][(len(acr[i][j]))-1]=='B':
                    mm=acr[i][j]
                    lab1.append(acr[i][j])
            lab.append(lab1.copy())
            lab1.clear()
        sub1=[]
        sub2=[]
        for i in range(len(acr)):
            for j in range(0,len(acr[i])):
                if acr[i][j][(len(acr[i][j]))-1]!='B':
                    sub1.append(acr[i][j])
            sub2.append(sub1.copy())
            sub1.clear()
        sub.clear()
        sub.extend(sub2)
        sub2.clear()

        w=Workbook()
        sheet=w.active
        l=['Mon','Tue','Wed','Thu','Fri']
        i1=2
        i2=2
        n=int(self.root.ids.count.text)#int(input("Enter the no of classes for time table: "))
        pp=int(self.root.ids.period.text)
        dd=int(self.root.ids.day.text)
        def table(i1,i2):
            for i in range(3,pp+3):
                sheet.cell(row=i1,column=i).value=i-2
            for i in range(i1,i1+dd):        
                sheet.cell(row=i+1,column=2).value=l[i-i2]

        for j in range(n) :
            table(i1,i2)
            i1=i1+8
            i2=i1
        t1=3
        t2=3

        comp=[]
        lim=[]   
        subb=[]
        def table_row(t1):
            vv=1
            for i in range (0,dd):
                for j in range (0,pp):
                    t1=3
                    for k in range(0,n):
                            if len(sub[k])>0:
                                
                                ch=ra.choice(sub[k])
                                if j<7:
                                    if k==0:
                                        sheet.cell(row=t1+i,column=j+3).value=ch
                                        up[ch]=int(up[ch])-1
                                        subb.append(ch)
                                    if k>0:
                                        if dic[ch]!=dic[subb[0]]:
                                            sheet.cell(row=t1+i,column=j+3).value=ch
                                            up[ch]=int(up[ch])-1
                                        while dic[ch]==dic[subb[0]]:
                                            ch=ra.choice(sub[k])
                                if up[ch]==0:
                                    sub[k].remove(ch)
                                t1=t1+8
                            
                            else: 
                                break
        lab3=[]
        ch_l=[0,1,2,3,4]                          
        def lab_ch():
            t1=3 
            for k in range(0,len(lab)):
                for jk in range(0,len(lab[k])): 
                    ch=ra.choice(ch_l) 
                    ch1_Lab=(lab[k][jk])  
                    for i in range (0,dd):
                        
                        for j in range (0,pp):
                            
                            if i==ch and j>3 and up[ch1_Lab]>0:
                                    for kk in range(up[ch1_Lab]):
                                        if sheet.cell(row=(t1+i),column=j+3+kk).value!=sheet.cell(row=(20),column=22).value:
                                            lab3.append(sheet.cell(row=(t1+i),column=j+3+kk).value)
                                            sheet.cell(row=t1+i,column=j+3+kk).value=ch1_Lab
                                            up[ch1_Lab]=int(up[ch1_Lab])-1
                                        else:
                                            sheet.cell(row=t1+i,column=j+3+kk).value=ch1_Lab
                                            up[ch1_Lab]=int(up[ch1_Lab])-1
                                        
                                    break
                                
                t1=t1+8
        
        def last_j():
                t1=3
                k=0
                for s in range(0,n):
                    for i in range (0,dd):
                            for j in range (0,pp):
                                if sheet.cell(row=(t1+i),column=j+3).value==sheet.cell(row=(20),column=22).value and k<len(lab3):
                                        sheet.cell(row=(t1+i),column=j+3).value=lab3[k]  
                                        k=k+1
                    t1=t1+8          
                                
        table_row(t1)
        lab_ch()
        last_j()
        #print(lab3)
        #print(up)
        nk=[]
        for i in range(100):
            nk.append(i)
        v=ra.choice(nk)
        w.save('Output/Time_table_'+str(v)+'.xlsx')


if __name__ == '__main__':
    AI_TimeTable().run()